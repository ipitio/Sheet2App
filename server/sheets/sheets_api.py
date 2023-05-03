from __future__ import print_function
import json
import os
import logging
from logging.handlers import RotatingFileHandler
from http import HTTPStatus

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from openpyxl.utils.cell import get_column_letter
from .utils import *
# from .auth import get_creds
from google.oauth2 import service_account

# Allow the API to have complete control over the spreadsheet with this scope
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Env variable for credentials.
# TODO: Migrate to dynamic token passed from end user request
CREDENTIALS = "sheets/service.json"


# Function to retrieve the token from local development environment. This is mainly for testing.
# TODO: remove this function to pull authentication token from request itself.
def get_creds():
    credentials = service_account.Credentials.from_service_account_file(
        CREDENTIALS, scopes=SCOPES)
    return credentials


# Log updates, errors, etc.
def setup_logger(app_id):
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    logger = logging.getLogger("app_%s" % app_id)
    logger.setLevel(logging.DEBUG)

    log_file = os.path.join(log_dir, "app_%s.log" % app_id)
    file_handler = RotatingFileHandler(log_file, maxBytes=10 * 1024 * 1024, backupCount=5)
    file_handler.setLevel(logging.INFO)

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    return logger


# Create a spreadsheet and return the new spreadsheet ID.
def create_spreadsheet(tokens, title, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Pass in the desired title of the spreadsheet into the call body.
        spreadsheet = {"properties": {"title": title}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )
        
        # Log the creation of the spreadsheet
        logger.info("Spreadsheet created: %s", spreadsheet.get("spreadsheetId"))

        # Return the newly created spreadsheets' ID number
        return spreadsheet.get("spreadsheetId"), HTTPStatus.OK
    except HttpError as err:
        print(err)
        logger.error(err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


# Retrieve data from a spreadsheet. If a range is specified, retrieve all data from within that range.
# If no range is specified, returns the entire first sheet. Reads the data row by row as a default.
def get_data(tokens, spreadsheet_id, sheet_id=None, range=None, majorDimension="ROWS", app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Call the Sheets API
        sheet = service.spreadsheets()

        # If no sheet_id is specified, then it is assumed that the first sheet is being requested.
        # If it is specified, then we need to find the name of the sheet with the specified sheet_id
        sheet_name = ""
        if sheet_id != None:
            sheets_info = (
                sheet.get(spreadsheetId=spreadsheet_id).execute().get("sheets")
            )

            for sheet_info in sheets_info:
                if sheet_info.get("properties").get("sheetId") == sheet_id:
                    sheet_name = sheet_info.get("properties").get("title")
                    break

        # If no sheet_id is specified, then default to the first sheet
        else:
            sheet_name = (
                sheet.get(spreadsheetId=spreadsheet_id)
                .execute()
                .get("sheets")[0]
                .get("properties")
                .get("title")
            )

        # Combine the sheet_name and range fields into a A1 notation range for sheets
        a1_range = ""
        if range is not None:
            a1_range = sheet_name + "!" + range
        else:
            a1_range = sheet_name

        # Make the API call to retrieve the specified data from the sheet
        result = (
            sheet.values()
            .get(
                spreadsheetId=spreadsheet_id,
                range=a1_range,
                majorDimension=majorDimension,
            )
            .execute()
        )
        
        logger.info("Data retrieved from spreadsheet: %s", spreadsheet_id)

        # Return the data stored in the spreadsheet as a 2 dimensional list.
        return result.get("values", []), HTTPStatus.OK
    except HttpError as err:
        print(err)
        logger.error(err)
        return [], HTTPStatus.INTERNAL_SERVER_ERROR


# Retrieve column data for certain specified columns Expects columns to be a
# array of column indices to retrieve the data for. Returns a 2 dimensional
# list containing the column data, where list[0] is the first column, etc...
def get_column_data(tokens, spreadsheet_id, sheet_id, columns, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Call the Sheets API
        sheet = service.spreadsheets()

        # If no sheet_id is specified, then it is assumed that the first sheet is being requested.
        # If it is specified, then we need to find the name of the sheet with the specified sheet_id
        sheet_name = ""
        if sheet_id is not None:
            sheets_info = (
                sheet.get(spreadsheetId=spreadsheet_id).execute().get("sheets")
            )

            for sheet_info in sheets_info:
                if sheet_info.get("properties").get("sheetId") == sheet_id:
                    sheet_name = sheet_info.get("properties").get("title")
                    break

        ranges = []
        for column in columns:
            column_letter = get_column_letter(column)
            ranges.append(sheet_name + "!" + column_letter + ":" + column_letter)

        # Make the API call to retrieve the specified data from the sheet
        result = (
            sheet.values()
            .batchGet(
                spreadsheetId=spreadsheet_id,
                ranges=ranges,
                majorDimension="COLUMNS",
            )
            .execute()
        )

        # Return the data stored in the spreadsheet as a 2 dimensional list, where each
        # entry in the list represents a column of data
        column_data = []
        for values in result.get("valueRanges"):
            column_data.extend(values.get("values"))
        
        logger.info("Column data retrieved from spreadsheet with ID: %s", spreadsheet_id)

        return column_data, HTTPStatus.OK
    except HttpError as err:
        print(err)
        logger.error(err)
        return [], HTTPStatus.INTERNAL_SERVER_ERROR


# Update a specific cell in the spreadsheet. This function only works for one
# cell in the spreadsheet, and does not support updating multiple cells at once.
def update_cell(tokens, spreadsheet_id, sheet_id, value_to_update, row_index, column_index, app_id=None):
    # Get the logger for the specific app
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Retrieve the type of value of value_to_update as a string parsable by the
        # Google Sheets API.
        type_of_value = get_spreadsheet_value_type(value_to_update)

        # Create the request body for the API call
        request_body = {
            "requests": [
                {
                    "updateCells": {
                        "start": {
                            "sheetId": sheet_id,
                            "rowIndex": row_index,
                            "columnIndex": column_index,
                        },
                        "rows": [
                            {
                                "values": [
                                    {
                                        "userEnteredValue": {
                                            type_of_value: value_to_update
                                        }
                                    }
                                ]
                            }
                        ],
                        "fields": "userEnteredValue",
                    }
                }
            ]
        }

        # Execute the API call
        res = (
            service.spreadsheets()
            .batchUpdate(spreadsheetId=spreadsheet_id, body=request_body)
            .execute()
        )
        
        # Log the update
        logger.info("Updated cell (%s, %s) in sheet %s of spreadsheet %s", row_index, column_index, sheet_id, spreadsheet_id)

        return res, HTTPStatus.OK

    except HttpError as err:
        print(err)
        logger.error("Error updating cell (%s, %s) in sheet %s of spreadsheet %s: %s", row_index, column_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


# For batch updating an entire record / row. Must pass the entire new row as an array
# to this API call.
def update_row(tokens, spreadsheet_id, sheet_id, updated_row_data, row_index, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Create the request body for the API call
        request_body = {
            "requests": [
                {
                    "updateCells": {
                        "start": {
                            "sheetId": sheet_id,
                            "rowIndex": row_index,
                            "columnIndex": 0,
                        },
                        "rows": [
                            {
                                "values": [
                                    {
                                        "userEnteredValue": {
                                            get_spreadsheet_value_type(value): value
                                        }
                                    }
                                    for value in updated_row_data
                                ]
                            }
                        ],
                        "fields": "userEnteredValue",
                    }
                }
            ]
        }

        # Execute the API call
        res = (
            service.spreadsheets()
            .batchUpdate(spreadsheetId=spreadsheet_id, body=request_body)
            .execute()
        )
        
        logger.info("Updated row %s in sheet %s of spreadsheet %s", row_index, sheet_id, spreadsheet_id)

        return res, HTTPStatus.OK

    except HttpError as err:
        print(err)
        logger.error("Error updating row %s in sheet %s of spreadsheet %s: %s", row_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


# Insert a new record / row in the spreadsheet. The data in row_to_insert must be passed as a list, where every element in the list
# corresponds to the elements in the row.
# TODO: Allow insertion of row into any row, not just end of spreadsheet
def insert_row(tokens, spreadsheet_id, sheet_id, row_to_insert, row_index=-1, app_id=None):
    logger = setup_logger(app_id)
    # If no row_index is specified, then the row is inserted at the end of the spreadsheet
    if row_index == -1:
        row_index = get_num_rows(tokens, spreadsheet_id, sheet_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Create the request body for the API call
        request_body = {
            "requests": [
                {
                    "insertDimension": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "ROWS",
                            "startIndex": row_index,
                            "endIndex": row_index + 1,
                        },
                        "inheritFromBefore": True,
                    }
                },
                {
                    "updateCells": {
                        "start": {
                            "sheetId": sheet_id,
                            "rowIndex": len(get_data(tokens, spreadsheet_id)[0]),
                            "columnIndex": 0,
                        },
                        "rows": [
                            {
                                "values": [
                                    {
                                        "userEnteredValue": {
                                            get_spreadsheet_value_type(value): value
                                        }
                                    }
                                    for value in row_to_insert
                                ]
                            }
                        ],
                        "fields": "userEnteredValue",
                    }
                },
            ]
        }

        # Execute the API call
        res = (
            service.spreadsheets()
            .batchUpdate(spreadsheetId=spreadsheet_id, body=request_body)
            .execute()
        )
        
        logger.info("Inserted row %s in sheet %s of spreadsheet %s", row_index, sheet_id, spreadsheet_id)

        return res, HTTPStatus.OK

    except HttpError as err:
        print(err)
        logger.error("Error inserting row %s in sheet %s of spreadsheet %s: %s", row_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


# Delete a record (or row) from the spreadsheet
def delete_row(tokens, spreadsheet_id, sheet_id, row_index, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Generate the request body
        request_body = {
            # An array of requests to send to the spreadsheet. We are only concerned with updating the sheet
            # in this API call.
            "requests": [
                {
                    "deleteRange": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_index,
                            "endRowIndex": row_index + 1,
                        },
                        # Shift all rows upwards if necessary
                        "shiftDimension": "ROWS",
                    }
                }
            ]
        }

        # Execute the API call
        res = (
            service.spreadsheets()
            .batchUpdate(spreadsheetId=spreadsheet_id, body=request_body)
            .execute()
        )
        
        logger.info("Deleted row %s in sheet %s of spreadsheet %s", row_index, sheet_id, spreadsheet_id)

        return res, HTTPStatus.OK

    except HttpError as err:
        print(err)
        logger.error("Error deleting row %s in sheet %s of spreadsheet %s: %s", row_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


# Write data to the first open column in the sheet
def write_column(tokens, spreadsheet_id, sheet_id, column_data, column_index, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build('sheets', 'v4', credentials=get_creds())
        
        sheet = service.spreadsheets()
        sheet_name = ""
        if sheet_id is not None:
            sheets_info = (
                sheet.get(spreadsheetId=spreadsheet_id).execute().get("sheets")
            )

            for sheet_info in sheets_info:
                if sheet_info.get("properties").get("sheetId") == sheet_id:
                    sheet_name = sheet_info.get("properties").get("title")
                    break
                
        column_letter = get_column_letter(column_index)
        update_range = f'{sheet_name}!{column_letter}1:{column_letter}{len(column_data)}'
        request_body = {
            'range': update_range,
            'values': [[value] for value in column_data],
        }
        request = sheet.values().update(
            spreadsheetId=spreadsheet_id,
            range=update_range,
            valueInputOption='RAW',
            body=request_body,
        )
        
        res = request.execute()
        
        logger.info("Wrote column %s in sheet %s of spreadsheet %s", column_index, sheet_id, spreadsheet_id)

        return res, HTTPStatus.OK
    
    except HttpError as err:
        print(err)
        logger.error("Error writing column %s in sheet %s of spreadsheet %s: %s", column_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR
    
    
def delete_column(tokens, spreadsheet_id, sheet_id, column_index, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build('sheets', 'v4', credentials=get_creds())
        
        sheet = service.spreadsheets()
        sheet_name = ""
        if sheet_id is not None:
            sheets_info = (
                sheet.get(spreadsheetId=spreadsheet_id).execute().get("sheets")
            )

            for sheet_info in sheets_info:
                if sheet_info.get("properties").get("sheetId") == sheet_id:
                    sheet_name = sheet_info.get("properties").get("title")
                    break
        
        column_letter = get_column_letter(column_index)
        update_range = f'{sheet_name}!{column_letter}:{column_letter}'
        
        request = sheet.values().clear(
            spreadsheetId=spreadsheet_id,
            range=update_range
        )
        
        res = request.execute()
        
        logger.info("Cleared column %s in sheet %s of spreadsheet %s", column_index, sheet_id, spreadsheet_id)
        
        return res, HTTPStatus.OK
    except HttpError as err:
        print(err)
        logger.error("Error deleting column %s in sheet %s of spreadsheet %s: %s", column_index, sheet_id, spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


def get_metadata(tokens, spreadsheet_id, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        sheet_metadata = (
            service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        )
        sheets = sheet_metadata.get("sheets", "")
        
        logger.info("Retrieved metadata for spreadsheet %s", spreadsheet_id)

        return sheets, HTTPStatus.OK
    except HttpError as err:
        print(err)
        logger.error("Error retrieving metadata for spreadsheet %s: %s", spreadsheet_id, err)
        return err, HTTPStatus.INTERNAL_SERVER_ERROR


def get_num_rows(tokens, spreadsheet_id, sheet_id, app_id=None):
    logger = setup_logger(app_id)
    try:
        service = build("sheets", "v4", credentials=get_creds())

        # Get the metadata of the spreadsheet
        sheet_metadata = (
            service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        )
        sheets = sheet_metadata.get("sheets", "")

        # Get the number of rows in the sheet
        for sheet in sheets:
            if sheet["properties"]["sheetId"] == sheet_id:
                num_rows = sheet["properties"]["gridProperties"]["rowCount"]
                logger.info("Retrieved number of rows for sheet %s of spreadsheet %s", sheet_id, spreadsheet_id)
                return num_rows
        logger.error("Error retrieving number of rows for sheet %s of spreadsheet %s", sheet_id, spreadsheet_id)
        return 0
    except HttpError as err:
        print(err)
        logger.error("Error retrieving number of rows for sheet %s of spreadsheet %s: %s", sheet_id, spreadsheet_id, err)
        return 0
    
    
def get_end_user_roles(tokens, role_mem_url, email, app_id=None):
    logger = setup_logger(app_id)
    spreadsheet_id = get_spreadsheet_id(role_mem_url)
    sheet_id = get_gid(role_mem_url)
    
    role_data, response_code = get_data(tokens=tokens, spreadsheet_id=spreadsheet_id, sheet_id=sheet_id, majorDimension="COLUMNS")
    if response_code != HTTPStatus.OK:
        logger.error("Error retrieving role data for sheet %s of spreadsheet %s: %s", sheet_id, spreadsheet_id, role_data)
        return [], response_code
    
    roles = []
    for role_col in role_data[1:]:
        if email in role_col[1:]:
            roles.append(role_col[0])
            
    logger.info("Retrieved roles for user %s in sheet %s of spreadsheet %s", email, sheet_id, spreadsheet_id)
            
    return roles, HTTPStatus.OK


def get_longest_column_length(tokens, spreadsheet_url, column_indexes):
    spreadsheet_id = get_spreadsheet_id(spreadsheet_url)
    sheet_id = get_gid(spreadsheet_url)
    
    data, response_code = get_data(tokens=tokens, spreadsheet_id=spreadsheet_id, sheet_id=sheet_id, majorDimension="COLUMNS")
    if response_code != HTTPStatus.OK:
        return -1, response_code
    
    longest = 0
    for index in column_indexes:
        longest = max(longest, len(data[index][1:]))
            
    return longest, HTTPStatus.OK
